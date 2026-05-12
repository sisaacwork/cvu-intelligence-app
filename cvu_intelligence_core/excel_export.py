"""Tier 1 — Excel Brief export.

Port of cvu_intelligence_generator.py::generate_tier1, refactored to
return bytes (for st.download_button) instead of writing to disk.

The Excel is a deliverable for staff to download; it never goes to
WordPress. Use alongside the WP Brief/Report drafts when staff want
an offline / shareable artifact of the underlying building data.
"""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .helpers import get_function, get_status


def generate_excel_brief(buildings, report_name, min_height=75):
    """Build the Tier 1 .xlsx in memory and return its bytes.

    buildings   — list of 20-tuples as returned by pull_mysql_data.
    report_name — display name shown in the title row.
    min_height  — included in the subtitle for context.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Buildings"

    date_str = datetime.now().strftime("%B %d, %Y")
    dark_fill = PatternFill(start_color="171717", end_color="171717", fill_type="solid")

    ws["A1"] = f"CVU Intelligence Brief — {report_name}"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="B4E817")
    ws["A1"].fill = dark_fill
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 24

    ws["A2"] = f"Buildings {min_height}m+ | Generated {date_str}"
    ws["A2"].font = Font(name="Arial", size=10, color="A0A0A0")
    ws["A2"].fill = dark_fill
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 18

    headers = ["Building Name", "City", "Height (m)", "Floors", "Year",
               "Status", "Function", "Material", "GFA (m²)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = h
        cell.font = Font(name="Arial", size=11, bold=True, color="B4E817")
        cell.fill = dark_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    sorted_b = sorted(buildings, key=lambda r: (r[3] or 0), reverse=True)
    for row_idx, b in enumerate(sorted_b, 5):
        row_color = "1E1E1E" if (row_idx - 5) % 2 == 0 else "252525"
        fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        vals = [
            b[1],
            b[2],
            f"{b[3]:.2f}" if b[3] else "—",
            b[4] or "—",
            str(b[5]) if b[5] else "—",
            get_status(b[7]),
            get_function(b),
            b[6] or "—",
            f"{b[13]:,.0f}" if b[13] else "—",
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = Font(name="Arial", size=10, color="FCFCFC")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

    widths = [35, 15, 12, 10, 8, 18, 30, 18, 14]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_filename(report_name):
    """Return a safe filename derived from the report name."""
    safe = report_name.replace("/", "-").replace("\\", "-")
    return f"CVU_Intelligence_Brief_{safe}.xlsx"
